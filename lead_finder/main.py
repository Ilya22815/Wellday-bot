import json
import os
from datetime import datetime
from email_finder import find_emails_on_site
from scorer import score
from letter_generator import generate_letter

# === Источники данных ===
USE_ZOON         = False   # заблокирован (недоступен с текущей сети)
USE_HH           = False   # API закрыт для анонимного доступа
USE_RUSPROFILE   = False   # реестр юрлиц, без контактов — бесполезен
USE_YELL         = True    # бизнес-справочник с телефонами и сайтами
YANDEX_API_KEY   = ""      # ключ Яндекс карт (опционально)

MIN_SCORE        = 30

SEEN_FILE = "seen.json"
CRM_FILE  = "crm.xlsx"

# Домены справочников — не ищем email на их страницах
DIRECTORY_DOMAINS = ["zoon.ru", "hh.ru", "rusprofile.ru", "yell.ru", "2gis.com"]


def load_seen():
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_seen(seen):
    with open(SEEN_FILE, "w", encoding="utf-8") as f:
        json.dump(list(seen), f, ensure_ascii=False, indent=2)


def collect_companies():
    companies = []
    seen_names = set()

    if USE_HH:
        from hh_parser import get_companies as hh_get
        print("Ищем через hh.ru...")
        hh_companies = hh_get()
        added = 0
        for c in hh_companies:
            key = c["name"].lower().strip()
            if key not in seen_names:
                seen_names.add(key)
                c["source"] = "hh.ru"
                companies.append(c)
                added += 1
        print(f"   hh.ru: {added} компаний")

    if USE_ZOON:
        from zoon_parser import get_companies as zoon_get
        print("Ищем через Zoon.ru...")
        zoon_companies = zoon_get()
        added = 0
        for c in zoon_companies:
            key = c["name"].lower().strip()
            if key not in seen_names:
                seen_names.add(key)
                c["source"] = "Zoon"
                companies.append(c)
                added += 1
        print(f"   Zoon: {added} новых компаний")

    if USE_YELL:
        from yell_parser import get_companies as yell_get
        print("Ищем через Yell.ru...")
        yell_companies = yell_get()
        added = 0
        for c in yell_companies:
            key = c["name"].lower().strip()
            if key not in seen_names:
                seen_names.add(key)
                c["source"] = "Yell"
                companies.append(c)
                added += 1
        print(f"   Yell: {added} новых компаний")

    if YANDEX_API_KEY:
        from yandex_parser import get_companies as yandex_get
        print("Ищем через Яндекс...")
        yandex_companies = yandex_get(api_key=YANDEX_API_KEY)
        added = 0
        for c in yandex_companies:
            key = c["name"].lower().strip()
            if key not in seen_names:
                seen_names.add(key)
                c["source"] = "Яндекс"
                companies.append(c)
                added += 1
        print(f"   Яндекс: {added} новых компаний")

    return companies


def update_crm(hot_leads):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    headers    = ["Компания", "Телефон", "Email / Сайт", "Отрасль", "Письмо отправлено", "Звонил", "Дата"]
    col_widths = [35, 18, 35, 25, 20, 12, 14]

    if os.path.exists(CRM_FILE):
        wb = openpyxl.load_workbook(CRM_FILE)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                existing.add(str(row[0]).lower().strip())
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CRM"
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[chr(64 + col)].width = width
        ws.row_dimensions[1].height = 20
        existing = set()

    added = 0
    for c in hot_leads:
        name = c.get("name", "")
        if name.lower().strip() in existing:
            continue

        emails = c.get("emails", [])
        site = c.get("site_url", "")
        contact = ", ".join(emails[:2]) if emails else (site or "—")

        row = [
            name,
            c.get("phone") or "—",
            contact,
            ", ".join(c.get("industries", [])[:2]),
            "Нет",
            "Нет",
            datetime.now().strftime("%d.%m.%Y"),
        ]
        ws.append(row)

        last_row = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[last_row].height = 30

        existing.add(name.lower().strip())
        added += 1

    wb.save(CRM_FILE)
    return added


def save_excel(hot_leads):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"

    headers    = ["Компания", "Score", "Отрасль", "Телефон", "Email / Сайт", "Адрес", "Письмо"]
    col_widths = [35, 8, 25, 20, 35, 35, 80]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + col)].width = width

    ws.row_dimensions[1].height = 20

    for row_num, c in enumerate(hot_leads, 2):
        emails = c.get("emails", [])
        site = c.get("site_url", "")
        contact = ", ".join(emails[:2]) if emails else (site or "—")

        values = [
            c.get("name"),
            c.get("score"),
            ", ".join(c.get("industries", [])),
            c.get("phone") or "—",
            contact,
            c.get("address") or "—",
            c.get("letter", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 80

    wb.save("leads.xlsx")


def run():
    seen = load_seen()
    print(f"Уже обработано ранее: {len(seen)} компаний")

    companies = collect_companies()
    print(f"Всего найдено: {len(companies)}")

    new_companies = [c for c in companies if c["name"].lower().strip() not in seen]
    print(f"Новых компаний: {len(new_companies)}\n")

    if not new_companies:
        print("Новых компаний нет — все уже были найдены ранее.")
        return

    hot_leads = []

    for i, company in enumerate(new_companies, 1):
        name = company.get("name", "")
        site = company.get("site_url", "")

        print(f"[{i}/{len(new_companies)}] {name}")

        # Не ищем email на страницах справочников
        if site and not any(d in site for d in DIRECTORY_DOMAINS):
            emails = find_emails_on_site(site)
        else:
            emails = []
        # Фильтруем email самих справочников
        emails = [e for e in emails if not any(f"@{d}" in e for d in DIRECTORY_DOMAINS)]
        company["emails"] = emails

        s = score(company)
        company["score"] = s

        phone = company.get("phone") or "нет"
        found = emails[:1] if emails else ("сайт: " + site if site else "не найдено")
        print(f"   score={s} | тел={phone} | email={found}")

        seen.add(name.lower().strip())

        if s >= MIN_SCORE:
            letter = generate_letter(company)
            company["letter"] = letter
            hot_leads.append(company)

    save_seen(seen)

    print(f"\nГорячих лидов: {len(hot_leads)}")

    if hot_leads:
        save_excel(hot_leads)
        crm_added = update_crm(hot_leads)
        print(f"Добавлено в CRM: {crm_added} компаний")
        print(f"Открой leads.xlsx — письма и детали")
        print(f"Открой crm.xlsx — твой трекер (письмо/звонок)")
    else:
        print("Горячих лидов не найдено.")


if __name__ == "__main__":
    run()
